"""
RBI Sectoral Bank Credit Ingestion

Downloads monthly sectoral deployment of bank credit data from RBI press releases.
Each month, RBI publishes an Excel file with sector-wise and industry-wise credit outstanding.

Process:
  1. Scrape rbi.org.in/Scripts/Data_Sectoral_Deployment.aspx for press release IDs
  2. Visit each press release page to find the Excel download URL
  3. Download Excel via requests (needs session cookies from rbi.org.in)
  4. Parse into flat records: (report_date, sector, sub_sector, outstanding_cr, yoy_pct)

Data goes back to ~Dec 2016.
"""

import re
import time
import requests
import openpyxl
import pandas as pd
from datetime import date, datetime
from pathlib import Path
from typing import Optional
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

BASE_DIR = Path(__file__).resolve().parents[2]
RAW_DIR = BASE_DIR / "data" / "raw" / "macro" / "rbi_credit"
CHROMEDRIVER = Path.home() / ".cache" / "selenium" / "chromedriver" / "mac-arm64" / "148.0.7778.167" / "chromedriver"

RBI_INDEX_URL = "https://rbi.org.in/Scripts/Data_Sectoral_Deployment.aspx"
RBI_PR_URL = "https://rbi.org.in/Scripts/BS_PressReleaseDisplay.aspx?prid={prid}"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.7778.168 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}


def _extract_report_date(title: str) -> Optional[date]:
    """Extract month/year from press release title like 'Sectoral Deployment of Bank Credit – March 2026'."""
    m = re.search(r"(\w+)\s+(\d{4})", title)
    if not m:
        return None
    month_name, year = m.group(1).lower(), int(m.group(2))
    month_num = MONTH_MAP.get(month_name)
    if not month_num:
        return None
    return date(year, month_num, 1)


def scrape_press_release_ids() -> list:
    """Use Selenium to scrape all press release IDs from the RBI sectoral deployment page."""
    from selenium.webdriver.chrome.options import Options as ChromeOptions

    opts = ChromeOptions()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")

    service = Service(executable_path=str(CHROMEDRIVER))
    driver = webdriver.Chrome(service=service, options=opts)
    driver.set_page_load_timeout(30)

    results = []
    try:
        driver.get(RBI_INDEX_URL)
        time.sleep(5)

        year_links = driver.find_elements(By.TAG_NAME, "a")
        years = []
        for l in year_links:
            t = l.text.strip()
            if t.isdigit() and 2017 <= int(t) <= 2030:
                years.append(t)

        for year in years:
            try:
                link = driver.find_element(By.LINK_TEXT, year)
                link.click()
                time.sleep(2)

                for l in driver.find_elements(By.PARTIAL_LINK_TEXT, "Sectoral"):
                    href = l.get_attribute("href") or ""
                    title = l.text.strip()
                    prid_match = re.search(r"prid=(\d+)", href)
                    if prid_match:
                        report_date = _extract_report_date(title)
                        if report_date:
                            results.append({
                                "prid": prid_match.group(1),
                                "title": title,
                                "report_date": report_date,
                            })
            except Exception:
                continue
    finally:
        driver.quit()

    results.sort(key=lambda x: x["report_date"])
    return results


def _find_excel_url(session: requests.Session, prid: str) -> Optional[str]:
    """Visit a press release page to find the Excel download link."""
    resp = session.get(RBI_PR_URL.format(prid=prid), timeout=20)
    if resp.status_code != 200:
        return None
    xlsx_match = re.search(r'href="(https://rbidocs\.rbi\.org\.in/[^"]+\.xlsx?)"', resp.text, re.IGNORECASE)
    if xlsx_match:
        return xlsx_match.group(1)
    xls_match = re.search(r'href="(https://rbidocs\.rbi\.org\.in/[^"]+\.xls)"', resp.text, re.IGNORECASE)
    return xls_match.group(1) if xls_match else None


def download_all_excels(start_date: Optional[date] = None) -> list:
    """Download all RBI sectoral credit Excel files."""
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    print("Scraping press release index...")
    releases = scrape_press_release_ids()
    print(f"Found {len(releases)} press releases")

    if start_date:
        releases = [r for r in releases if r["report_date"] >= start_date]
        print(f"  Filtered to {len(releases)} after {start_date}")

    session = requests.Session()
    session.headers.update(HEADERS)
    session.get("https://rbi.org.in/", timeout=15)
    time.sleep(1)

    downloaded = []
    for r in releases:
        out_path = RAW_DIR / f"rbi_credit_{r['report_date'].isoformat()}.xlsx"
        if out_path.exists():
            downloaded.append((r["report_date"], out_path))
            continue

        excel_url = _find_excel_url(session, r["prid"])
        if not excel_url:
            print(f"  {r['report_date']}: no Excel link found")
            continue

        resp = session.get(excel_url, timeout=30)
        if resp.status_code == 200 and resp.content[:4] in (b"PK\x03\x04", b"\xd0\xcf\x11\xe0"):
            out_path.write_bytes(resp.content)
            downloaded.append((r["report_date"], out_path))
            print(f"  {r['report_date']}: downloaded ({len(resp.content) // 1024}KB)")
        else:
            print(f"  {r['report_date']}: download failed (HTTP {resp.status_code})")

        time.sleep(0.5)

    return downloaded


def parse_rbi_credit_excel(path: Path, report_date: date) -> pd.DataFrame:
    """Parse an RBI sectoral credit Excel into flat records.

    Layout: col0=sector, col1..col3=outstanding (oldest to newest), col4..col5=YoY% (oldest to newest).
    The latest outstanding is the last numeric col before the % columns, and latest YoY% is the last col.
    Column count can vary across months, so we detect dynamically.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        import xlrd
        return _parse_xls(path, report_date)

    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        statement_type = "sector" if "1" in sheet_name else "industry"

        # Find the header row with "%" to determine column layout
        pct_cols = []
        outstanding_cols = []
        for r_idx in range(1, min(10, ws.max_row + 1)):
            row_vals = [ws.cell(r_idx, c).value for c in range(1, ws.max_column + 1)]
            if "%" in [str(v).strip() for v in row_vals if v]:
                for c_idx, v in enumerate(row_vals):
                    if str(v).strip() == "%":
                        pct_cols.append(c_idx)

        # Default: last column is YoY%, column before pct_cols is latest outstanding
        last_col = ws.max_column - 1
        yoy_col = last_col  # last column (0-indexed)
        outstanding_col = pct_cols[0] - 1 if pct_cols else last_col - 2  # last outstanding before % cols

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            label = row[0].value
            if not label or not isinstance(label, str):
                continue
            label_stripped = label.strip()
            if not label_stripped or label_stripped.startswith("(") or label_stripped in ("Sector", "Industry", "Priority Sector (Memo)"):
                continue

            outstanding = _safe_float(row[outstanding_col].value if outstanding_col < len(row) else None)
            yoy_pct = _safe_float(row[yoy_col].value if yoy_col < len(row) else None)

            if outstanding is None:
                continue
            # YoY should be a percentage (-100 to +500 ish), not a large number
            if yoy_pct is not None and abs(yoy_pct) > 1000:
                yoy_pct = None

            depth = 0
            if label.startswith("    "):
                depth = 2
            elif label.startswith("  "):
                depth = 1

            records.append({
                "report_date": report_date.isoformat(),
                "statement": statement_type,
                "sector": label_stripped,
                "depth": depth,
                "outstanding_cr": round(outstanding, 2),
                "yoy_pct": round(yoy_pct, 2) if yoy_pct else None,
            })

    return pd.DataFrame(records)


def _parse_xls(path: Path, report_date: date) -> pd.DataFrame:
    """Fallback parser for old .xls format."""
    import xlrd
    wb = xlrd.open_workbook(str(path))
    records = []
    for sheet in wb.sheets():
        statement_type = "sector" if "1" in sheet.name else "industry"
        for rx in range(sheet.nrows):
            row = sheet.row(rx)
            label = row[0].value if row else None
            if not label or not isinstance(label, str):
                continue
            label = label.strip()
            if not label or label.startswith("("):
                continue
            outstanding = _safe_float(row[3].value if len(row) > 3 else None)
            if outstanding is None:
                continue
            yoy_pct = _safe_float(row[5].value if len(row) > 5 else None)
            records.append({
                "report_date": report_date.isoformat(),
                "statement": statement_type,
                "sector": label,
                "depth": 0,
                "outstanding_cr": round(outstanding, 2),
                "yoy_pct": round(yoy_pct, 2) if yoy_pct else None,
            })
    return pd.DataFrame(records)


def _safe_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        f = float(val)
        return f if f != 0 else None
    except (ValueError, TypeError):
        return None


def download_and_parse_all(start_date: Optional[date] = None) -> pd.DataFrame:
    """Download all Excel files and parse into a single DataFrame."""
    files = download_all_excels(start_date)
    print(f"\nParsing {len(files)} files...")

    all_dfs = []
    for report_dt, path in files:
        df = parse_rbi_credit_excel(path, report_dt)
        if not df.empty:
            all_dfs.append(df)

    if not all_dfs:
        return pd.DataFrame()

    combined = pd.concat(all_dfs, ignore_index=True)
    combined["report_date"] = pd.to_datetime(combined["report_date"])
    combined = combined.sort_values(["report_date", "statement", "sector"])
    print(f"Total RBI credit records: {len(combined)}")
    return combined


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Download RBI Sectoral Credit Data")
    parser.add_argument("--start", type=str, help="Start date YYYY-MM-DD")
    args = parser.parse_args()

    start = date.fromisoformat(args.start) if args.start else None
    df = download_and_parse_all(start)
    if not df.empty:
        print(f"\nDate range: {df['report_date'].min()} to {df['report_date'].max()}")
        print(f"Months: {df['report_date'].nunique()}")
        print(df.head(20))
