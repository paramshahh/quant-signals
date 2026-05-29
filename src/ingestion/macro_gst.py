"""
GST Collections Ingestion

Downloads state-wise monthly GST collection data from tutorial.gst.gov.in.
Excel files are published per fiscal year with CGST/SGST/IGST breakdowns by state.
Available from FY 2017-18 onward.
"""

import requests
import openpyxl
import pandas as pd
from datetime import date
from pathlib import Path
from typing import Optional

BASE_DIR = Path(__file__).resolve().parents[2]
RAW_DIR = BASE_DIR / "data" / "raw" / "macro" / "gst"

GST_URL = "https://tutorial.gst.gov.in/offlineutilities/gst_statistics/statewise_GST_collection_{fy}.xlsx"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
}


def _fy_label(start_year: int) -> str:
    return f"{start_year}-{(start_year + 1) % 100:02d}"


def download_gst_excel(fy_start: int) -> Optional[Path]:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    fy = _fy_label(fy_start)
    url = GST_URL.format(fy=fy)
    out_path = RAW_DIR / f"statewise_GST_collection_{fy}.xlsx"

    if out_path.exists():
        print(f"  Already have {out_path.name}")
        return out_path

    resp = requests.get(url, headers=HEADERS, timeout=30)
    if resp.status_code != 200:
        print(f"  FY {fy}: HTTP {resp.status_code}")
        return None

    out_path.write_bytes(resp.content)
    print(f"  Downloaded {out_path.name} ({len(resp.content) // 1024}KB)")
    return out_path


def parse_gst_excel(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)

    if "Collections-Statewise" in wb.sheetnames:
        ws = wb["Collections-Statewise"]
        header_row = 5
        sub_header_row = 6
        data_start_row = 7
    else:
        ws = wb[wb.sheetnames[0]]
        header_row = 5
        sub_header_row = 6
        data_start_row = 7

    months = []
    header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False))[0]
    for cell in header_cells:
        if cell.value and hasattr(cell.value, "strftime"):
            months.append((cell.column, cell.value))

    records = []
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, values_only=False):
        state_cd = row[0].value
        state_name = row[1].value
        if not state_name or not isinstance(state_cd, (int, float)):
            continue

        for month_col, month_dt in months:
            col_idx = month_col - 1
            cgst = row[col_idx].value if col_idx < len(row) else None
            sgst = row[col_idx + 1].value if col_idx + 1 < len(row) else None
            igst = row[col_idx + 2].value if col_idx + 2 < len(row) else None
            total = row[col_idx + 3].value if col_idx + 3 < len(row) else None

            if cgst is None and sgst is None and igst is None:
                continue

            records.append({
                "month": month_dt.strftime("%Y-%m-01"),
                "state_code": int(state_cd),
                "state": state_name.strip(),
                "cgst_cr": _to_float(cgst),
                "sgst_cr": _to_float(sgst),
                "igst_cr": _to_float(igst),
                "total_cr": _to_float(total),
            })

    return pd.DataFrame(records)


def _to_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return None


def download_and_parse_all(start_fy: int = 2017, end_fy: int = 2025) -> pd.DataFrame:
    all_dfs = []
    for fy_start in range(start_fy, end_fy + 1):
        path = download_gst_excel(fy_start)
        if path:
            df = parse_gst_excel(path)
            if not df.empty:
                print(f"  FY {_fy_label(fy_start)}: {len(df)} records")
                all_dfs.append(df)

    if not all_dfs:
        return pd.DataFrame()

    combined = pd.concat(all_dfs, ignore_index=True)
    combined["month"] = pd.to_datetime(combined["month"])
    combined = combined.drop_duplicates(subset=["month", "state_code"]).sort_values(["month", "state_code"])
    print(f"\nTotal GST records: {len(combined)}")
    return combined


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Download GST Collections")
    parser.add_argument("--start-fy", type=int, default=2017, help="Start fiscal year (e.g. 2017 for FY 2017-18)")
    parser.add_argument("--end-fy", type=int, default=2025, help="End fiscal year (e.g. 2025 for FY 2025-26)")
    args = parser.parse_args()

    df = download_and_parse_all(args.start_fy, args.end_fy)
    if not df.empty:
        print(f"\n{df.head(10)}")
        print(f"\nDate range: {df['month'].min()} to {df['month'].max()}")
        print(f"States: {df['state'].nunique()}")
